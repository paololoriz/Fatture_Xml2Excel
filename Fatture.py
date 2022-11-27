from datetime import datetime
import os, glob
import xmltodict
import calendar
import openpyxl
from openpyxl.utils import get_column_letter
import string
import tkinter
from tkinter import filedialog
from openpyxl.styles import PatternFill

tkinter.Tk().withdraw()
path = filedialog.askdirectory() #Scegliere cartella da cui importare fatture

DatiFatture = []
StrutturaCampi = []
alphabet = list(string.ascii_uppercase)

# Creiamo una struttura dei campi che ci interessa esportare seguendo la struttura xml, se i campi sono composti da più righe,
# come nel caso degli articoli, prendiamo il primo campo 'padre' necessario (DettaglioLinee) il programma si occuperà di prendere
# tutti i sottocampi contenuti.

StrutturaCampi = [['FatturaElettronicaBody', 'DatiGenerali', 'DatiGeneraliDocumento', 'Data'],
                  ['FatturaElettronicaBody', 'DatiGenerali', 'DatiGeneraliDocumento', 'Numero'],
                  ['FatturaElettronicaBody', 'DatiGenerali', 'DatiGeneraliDocumento', 'ImportoTotaleDocumento'],
                  ['FatturaElettronicaHeader', 'CessionarioCommittente', 'DatiAnagrafici', 'Anagrafica', 'Denominazione'],
                  ['FatturaElettronicaBody', 'DatiPagamento'],
                  ['FatturaElettronicaBody', 'DatiBeniServizi', 'DettaglioLinee']]

# Qui indichiamo la certella con i file XML da elaborare
#path = "C:\\Users\\ricca\\Downloads\\FatturaE"

# Leggiamo tutti i file
for filename in glob.glob(os.path.join(path, "*.xml")):
    Fattura = []

    with open(filename) as fd:
        doc = xmltodict.parse(fd.read())
        # Prendiamo la chiave radice principale in automatico (es.: p:FatturaElettronica or ns1:FatturaElettronica)
        root = list(doc.keys())[0]

        # Facciamo il parse della struttura campi
        for fieldpath in StrutturaCampi:
            fieldname = ''

            # Prepariamo la stringa da interrogare successivamente con il metodo eval
            string_2_eval = 'doc["' + root + '"]'
            row_dict = {}

            for field in fieldpath:
                # memorizziamo il nome dell'ultimo campo della lista (Es.ProgressivoInvio):
                fieldname = fieldpath[-1]
                # accodiamo il percorso per andare a recuperare il valore nell'xml
                string_2_eval += '["' + field.replace('\n', '') + '"]'

            # Verifichiamo se il valore esiste e che non dia errori
            try:
                value = eval(string_2_eval)
            except:
                value = False

            # Verifichiamo che tipo di dato abbiamo:
            # se è un dictionary vuol dire che è una fattura con una sola riga di dettaglio
            if isinstance(value, dict) and value:
                for k, v in value.items():
                    row_dict[k] = v
                Fattura.append(('Righe fattura', row_dict))

            # se è una lista vuol dire che è una fattura con una più righe di dettaglio
            elif isinstance(value, list) and value:
                for l in value:
                    for k, v in l.items():
                        row_dict[k] = v
                    Fattura.append(('Righe fattura', row_dict))
                    row_dict = {}
            # se è un valore vuol dire che è un valore singolo
            elif not isinstance(value, dict) and value:
                Fattura.append((fieldname, value))
            # se è un valore ma non contiene dati aggiungiamo comunque il campo vuoto per non perdere la formattazione delle colonne nel file in uscita
            elif not isinstance(value, dict) and not value:
                Fattura.append((fieldname, ''))

        # Fine lettura campi fattura
        # A questo punto in Fattura[] abbiamo la lista di tutti i campi che ci interessano, dove ogni valore di 'Righe Fattura' conterrà
        # il dictionary con i campi della riga.

        # Volendo possiamo accodare a un'altra lista che conterrà tutte le fatture in formato lista.
        DatiFatture.append(Fattura)

for fattura in DatiFatture:
    if(fattura[1][1] == ''):
        DatiFatture.remove(fattura)


data = DatiFatture[0][0][1]
year = data[0:4]
month = data[5:7]
month = calendar.month_name[int(month)]
strpath = f'{month}_{year}.xlsx'
file_excel =openpyxl.Workbook()
sheet = file_excel.active
sheet.title = month

#sheet['A1'] = 'N'
#sheet['B1'] = 'IMPORTO TOT. FATTURA'
#sheet['C1'] = 'DATA'
#sheet['D1'] = 'FATTURA N'
#sheet['E1'] = ''
#sheet['F1'] = ''
#sheet['G1'] = 'DENOMINAZIONE'
#sheet['H1'] = ''
#sheet['I1'] = 'MODALITA\' PAGAMENTO'
#sheet['J1'] = 'DATA'
#sheet['K1'] = 'IMPORTO'

for fattura in DatiFatture:
    counter = DatiFatture.index(fattura)
    array_mod = []
    pagamento = ''
    descrizione = ''
    n_fattura = fattura[1][1]
    sheet['A'+str(counter+2)] =(int(n_fattura[5:10]))
    sheet['C'+str(counter+2)] = datetime.strptime(fattura[0][1], '%Y-%m-%d')
    sheet['D'+str(counter+2)] = fattura[1][1]
    sheet['E'+str(counter+2)] = float(fattura[2][1])
    sheet['F'+str(counter+2)] = '=B'+str(counter+2)+'-E'+str(counter+2) #float(fattura[2][1])-float(fattura[2][1])
    sheet['G'+str(counter+2)] = fattura[3][1]
    sheet['H'+str(counter+2)] = 'Fattura'

    linee = len(fattura)-5
    index = 1
    offset = 0
    importoTotPagamenti: float = 0
    importoTotFattura: float = 0
    #print(fattura[4][1])
    #print(fattura[1][1])
    if fattura[4][1] == '':
        sheet['A' + str(counter + 2)].fill = PatternFill(fgColor="FF0000", fill_type='solid')
    else:
        if isinstance(fattura[4][1].get('DettaglioPagamento'), list):
            for elem in fattura[4][1].get('DettaglioPagamento'):
                # for elem in fattura[4][1].get('DettaglioPagamento'):
                cod_pagamento = elem['ModalitaPagamento']
                match cod_pagamento:
                    case 'MP01':
                        mod_pagamento = 'Contanti'
                    case 'MP02':
                        mod_pagamento = 'Assegno'
                    case 'MP03':
                        mod_pagamento = 'Assegno circolare'
                    case 'MP04':
                        mod_pagamento = 'Contanti presso Tesoreria'
                    case 'MP05':
                        mod_pagamento = 'Bonifico'
                    case 'MP06':
                        mod_pagamento = 'Vaglia cambiario'
                    case 'MP07':
                        mod_pagamento = 'Bollettino bancario'
                    case 'MP08':
                        mod_pagamento = 'Carta di credito'
                    case '':
                        mod_pagamento = ''
                    case null:
                        mod_pagamento = ''
                array_mod.append(mod_pagamento)
        else:
            cod_pagamento = fattura[4][1]['DettaglioPagamento']['ModalitaPagamento']
            match cod_pagamento:
                case 'MP01':
                    mod_pagamento = 'Contanti'
                case 'MP02':
                    mod_pagamento = 'Assegno'
                case 'MP03':
                    mod_pagamento = 'Assegno circolare'
                case 'MP04':
                    mod_pagamento = 'Contanti presso Tesoreria'
                case 'MP05':
                    mod_pagamento = 'Bonifico'
                case 'MP06':
                    mod_pagamento = 'Vaglia cambiario'
                case 'MP07':
                    mod_pagamento = 'Bollettino bancario'
                case 'MP08':
                    mod_pagamento = 'Carta di credito'
                case '':
                    mod_pagamento = ''
                case null:
                    mod_pagamento = ''
            array_mod.append(mod_pagamento)
        pagamento = array_mod[0]
        while (index <= linee):
            sheet[get_column_letter(offset + 9) + str(counter + 2)] = pagamento
            descrizione = fattura[index + 4][1]['Descrizione']
            datarif = descrizione[len(descrizione) - 19:]
            datarif = datarif[0:10]
            sheet[get_column_letter(offset + 10) + str(counter + 2)] = datetime.strptime(datarif, '%d/%m/%Y')
            sheet[get_column_letter(offset + 11) + str(counter + 2)] = round(
                float(fattura[index + 4][1].get('PrezzoTotale')) +
                (float(fattura[index + 4][1].get('PrezzoTotale')) * 0.22), 2
            )
            importoTotFattura = importoTotFattura + round(
                float(fattura[index + 4][1].get('PrezzoTotale')) +
                (float(fattura[index + 4][1].get('PrezzoTotale')) * 0.22), 2
            )
            sheet['B' + str(counter + 2)] = importoTotFattura
            if len(array_mod) > 1 and len(array_mod) == len(set(array_mod)):
                sheet['A' + str(counter + 2)].fill = PatternFill(fgColor="FFA500", fill_type='solid')
            offset = offset + 3
            index = index + 1
ws = file_excel[month]
ws.delete_rows(len(DatiFatture)+2,ws.max_row)
file_excel.save(strpath)

# if isinstance(fattura[4][1].get('DettaglioPagamento'), list):    #Se ci sono più pagamenti
#     for elem in fattura[4][1].get('DettaglioPagamento'):
#         cod_pagamento = elem.get('ModalitaPagamento')
#         match cod_pagamento:
#             case 'MP01':
#                 mod_pagamento = 'Contanti'
#             case 'MP02':
#                 mod_pagamento = 'Assegno'
#             case 'MP03':
#                 mod_pagamento = 'Assegno circolare'
#             case 'MP04':
#                 mod_pagamento = 'Contanti presso Tesoreria'
#             case 'MP05':
#                 mod_pagamento = 'Bonifico'
#             case 'MP06':
#                 mod_pagamento = 'Vaglia cambiario'
#             case 'MP07':
#                 mod_pagamento = 'Bollettino bancario'
#             case 'MP08':
#                 mod_pagamento = 'Carta di credito'
#             case '':
#                 mod_pagamento = ''
#             case null:
#                 mod_pagamento = ''
#
#         array_mod.append(mod_pagamento)
#         sheet[get_column_letter(offset + 9) + str(counter + 2)] = mod_pagamento
#         sheet[get_column_letter(offset + 10) + str(counter + 2)] = datetime.strptime(fattura[0][1], '%Y-%m-%d')
#         sheet[get_column_letter(offset + 11) + str(counter + 2)] = float(elem.get('ImportoPagamento'))
#         importoTotPagamenti = importoTotPagamenti + float(elem.get('ImportoPagamento'))
#         sheet['B' + str(counter + 2)] = importoTotPagamenti
#
#         if  mod_pagamento[index] != mod_pagamento[index+1]:
#             sheet['A' + str(counter + 2)].fill = PatternFill(fgColor="FFA500", fill_type='solid')
#         offset = offset + 3
#         index = index + 1
#
# else:   #Se il pagamento è unico
#     cod_pagamento = fattura[4][1]['DettaglioPagamento']['ModalitaPagamento']
#     match cod_pagamento:
#         case 'MP01':
#             mod_pagamento = 'Contanti'
#         case 'MP02':
#             mod_pagamento = 'Assegno'
#         case 'MP03':
#             mod_pagamento = 'Assegno circolare'
#         case 'MP04':
#             mod_pagamento = 'Contanti presso Tesoreria'
#         case 'MP05':
#             mod_pagamento = 'Bonifico'
#         case 'MP06':
#             mod_pagamento = 'Vaglia cambiario'
#         case 'MP07':
#             mod_pagamento = 'Bollettino bancario'
#         case 'MP08':
#             mod_pagamento = 'Carta di credito'
#         case '':
#             mod_pagamento = ''
#         case null:
#             mod_pagamento = ''
#     sheet[get_column_letter(offset + 9) + str(counter + 2)] = mod_pagamento
#     sheet[get_column_letter(offset + 10) + str(counter + 2)] = datetime.strptime(fattura[0][1], '%Y-%m-%d')
#     sheet[get_column_letter(offset + 11) + str(counter + 2)] = float(fattura[4][1]['DettaglioPagamento']['ImportoPagamento'])
#     sheet['B' + str(counter + 2)] = float(fattura[4][1]['DettaglioPagamento']['ImportoPagamento'])
#     offset = offset + 3
#     index = index + 1

# if sheet['B' + str(counter + 2)] == sheet['E' + str(counter + 2)]:
#     sheet['B' + str(counter + 2)].fill = PatternFill(fgColor="4CBB17", fill_type='solid')
# else:
#     sheet['B' + str(counter + 2)].fill = PatternFill(fgColor="FF3701", fill_type='solid')

# if len(fattura) > 6 or mod_pagamento == '':
#     sheet['A' + str(counter + 2)].fill = PatternFill(fgColor="FFA500", fill_type='solid')
